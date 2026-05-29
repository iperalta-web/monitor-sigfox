#!/usr/bin/env python3
# coding: utf-8
"""
Análisis de Mensajes Sigfox por Dispositivo
Genera un Excel con:
  - Hoja "Datos": frameCount por dispositivo por día (tabla cruda)
  - Hoja "Resumen": totales, promedios, % del límite por dispositivo
  - Hoja "Config": parámetros usados
Uso:
  python analisis_mensajes.py [año] [mes]
  Ejemplos:
    python analisis_mensajes.py          # mes actual
    python analisis_mensajes.py 2025 9   # septiembre 2025
"""

import json
import os
import sys
import calendar
from datetime import date

import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
OK_FILL     = PatternFill("solid", fgColor="E8F5E9")
ADV_FILL    = PatternFill("solid", fgColor="FFF8D6")
CRIT_FILL   = PatternFill("solid", fgColor="FFE0E0")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def cargar_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def limite_dispositivo(cfg, device_id):
    overrides = cfg["limites"].get("por_dispositivo", {})
    return overrides.get(str(device_id), cfg["limites"]["diario_default"])


def consultar_consumo(login, password, device_id, year, month):
    url = f"https://api.sigfox.com/v2/devices/{device_id}/consumption/{year}/{month}"
    r = requests.get(url, auth=HTTPBasicAuth(login, password), timeout=15)
    if r.status_code == 200:
        return r.json().get("consumption", {}).get("consumptions", [])
    print(f"  [WARN] {device_id}: HTTP {r.status_code}")
    return None


def recopilar_datos(cfg, ids, year, month):
    login    = cfg["sigfox"]["login"]
    password = cfg["sigfox"]["password"]
    dias_mes = calendar.monthrange(int(year), int(month))[1]

    registros = []
    for i, device_id in enumerate(ids):
        print(f"  [{i+1}/{len(ids)}] Consultando {device_id}...")
        consumptions = consultar_consumo(login, password, device_id, year, month)
        if consumptions is None:
            continue

        fila = {"ID Dispositivo": device_id}
        total = 0
        for d in range(1, dias_mes + 1):
            idx = d - 1
            val = consumptions[idx]["frameCount"] if idx < len(consumptions) else "NA"
            fila[str(d)] = val
            if isinstance(val, (int, float)):
                total += val
        fila["TOTAL"] = total
        registros.append(fila)

    return registros, dias_mes


def estilo_celda(ws, cell, fill=None, font=None, alignment=None, border=True, fmt=None):
    if fill:      cell.fill = fill
    if font:      cell.font = font
    if alignment: cell.alignment = alignment
    if border:    cell.border = BORDER
    if fmt:       cell.number_format = fmt


def escribir_hoja_datos(wb, registros, dias_mes, year, month, cfg):
    ws = wb.create_sheet("Datos por Día")
    dias_col = [str(d) for d in range(1, dias_mes + 1)]
    headers = ["ID Dispositivo"] + dias_col + ["TOTAL"]

    # Fila de encabezado de mes/año
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    titulo = ws.cell(1, 1, f"Consumo de Mensajes Sigfox — {calendar.month_name[int(month)]} {year}")
    titulo.font = Font(bold=True, name="Arial", size=12, color="1A73E8")
    titulo.alignment = CENTER

    # Encabezados
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        estilo_celda(ws, c, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    # Datos
    limite_default = cfg["limites"]["diario_default"]
    for row, rec in enumerate(registros, 3):
        device_id = rec["ID Dispositivo"]
        limite = limite_dispositivo(cfg, device_id)
        total = rec["TOTAL"]
        pct = total / (limite * dias_mes) * 100 if limite > 0 else 0
        row_fill = CRIT_FILL if pct >= cfg["alertas"]["umbral_critico_pct"] else \
                   ADV_FILL  if pct >= cfg["alertas"]["umbral_advertencia_pct"] else OK_FILL

        for col, h in enumerate(headers, 1):
            val = rec.get(h, "NA")
            c = ws.cell(row, col)
            c.value = val
            c.font = DATA_FONT
            c.border = BORDER
            if col == 1:
                c.alignment = LEFT
            elif h == "TOTAL":
                c.alignment = RIGHT
                c.fill = row_fill
                c.font = Font(name="Arial", size=10, bold=True)
            else:
                c.alignment = CENTER
                if isinstance(val, (int, float)) and val > limite:
                    c.fill = PatternFill("solid", fgColor="FFB3B3")
                elif isinstance(val, (int, float)):
                    c.fill = PatternFill("solid", fgColor="FAFAFA")

    # Anchos
    ws.column_dimensions["A"].width = 22
    for d in range(1, dias_mes + 1):
        ws.column_dimensions[get_column_letter(d + 1)].width = 6
    ws.column_dimensions[get_column_letter(dias_mes + 2)].width = 10
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "B3"
    return ws


def escribir_hoja_resumen(wb, registros, dias_mes, year, month, cfg, dia_corte=None):
    ws = wb.create_sheet("Resumen")
    if dia_corte is None:
        dia_corte = dias_mes

    headers = [
        "ID Dispositivo", "Total mensajes", "Límite/día",
        "Límite mes completo", "Límite al día de corte",
        "% del límite (mes)", "% del límite (al corte)",
        "Promedio/día", "Días con datos", "Estado"
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    titulo = ws.cell(1, 1, f"Resumen — {calendar.month_name[int(month)]} {year}  (corte día {dia_corte})")
    titulo.font = Font(bold=True, name="Arial", size=12, color="1A73E8")
    titulo.alignment = CENTER

    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        estilo_celda(ws, c, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    for row, rec in enumerate(registros, 3):
        device_id  = rec["ID Dispositivo"]
        total      = rec["TOTAL"]
        limite_dia = limite_dispositivo(cfg, device_id)
        limite_mes = limite_dia * dias_mes
        limite_corte = limite_dia * dia_corte
        dias_datos = sum(1 for d in range(1, dias_mes + 1)
                         if isinstance(rec.get(str(d)), (int, float)))
        promedio   = round(total / dias_datos, 1) if dias_datos > 0 else 0
        pct_mes    = round(total / limite_mes * 100, 1) if limite_mes > 0 else 0
        pct_corte  = round(total / limite_corte * 100, 1) if limite_corte > 0 else 0

        umbral_crit = cfg["alertas"]["umbral_critico_pct"]
        umbral_adv  = cfg["alertas"]["umbral_advertencia_pct"]
        if pct_corte >= umbral_crit:
            estado = "CRÍTICO"; rfill = CRIT_FILL
        elif pct_corte >= umbral_adv:
            estado = "ADVERTENCIA"; rfill = ADV_FILL
        else:
            estado = "OK"; rfill = OK_FILL

        vals = [device_id, total, limite_dia, limite_mes, limite_corte,
                pct_mes / 100, pct_corte / 100, promedio, dias_datos, estado]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, val)
            c.font = DATA_FONT
            c.border = BORDER
            if col in (6, 7):
                c.number_format = "0.0%"
                c.alignment = RIGHT
                c.fill = rfill
            elif col == 10:
                c.alignment = CENTER
                c.fill = rfill
                c.font = Font(name="Arial", size=10, bold=True)
            elif col == 1:
                c.alignment = LEFT
            else:
                c.alignment = RIGHT

    ws.column_dimensions["A"].width = 22
    for col, w in zip("BCDEFGHIJ", [16, 12, 18, 20, 18, 18, 14, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B3"
    return ws


def escribir_hoja_config(wb, cfg, year, month):
    ws = wb.create_sheet("Configuración")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    params = [
        ("Parámetro", "Valor"),
        ("Año consultado", year),
        ("Mes consultado", month),
        ("Login Sigfox", cfg["sigfox"]["login"]),
        ("Archivo dispositivos", cfg["archivos"]["lista_devices"]),
        ("Límite diario default", cfg["limites"]["diario_default"]),
        ("Umbral advertencia (%)", cfg["alertas"]["umbral_advertencia_pct"]),
        ("Umbral crítico (%)", cfg["alertas"]["umbral_critico_pct"]),
        ("Email habilitado", cfg["alertas"]["email"]["habilitado"]),
        ("Destinatarios", ", ".join(cfg["alertas"]["email"]["destinatarios"])),
    ]
    overrides = cfg["limites"].get("por_dispositivo", {})
    if overrides:
        params.append(("", ""))
        params.append(("Límites individuales", ""))
        for dev, lim in overrides.items():
            params.append((f"  {dev}", lim))

    for row, (k, v) in enumerate(params, 1):
        ck = ws.cell(row, 1, k)
        cv = ws.cell(row, 2, v)
        if row == 1:
            ck.fill = HEADER_FILL; ck.font = HEADER_FONT
            cv.fill = HEADER_FILL; cv.font = HEADER_FONT
        ck.border = BORDER; cv.border = BORDER
        ck.font = Font(bold=(row == 1), name="Arial", size=10)
        cv.font = Font(name="Arial", size=10)
    return ws


def main():
    hoy = date.today()
    year  = str(sys.argv[1]) if len(sys.argv) > 1 else str(hoy.year)
    month = str(sys.argv[2]) if len(sys.argv) > 2 else str(hoy.month)
    dia_corte = hoy.day if (int(year) == hoy.year and int(month) == hoy.month) else None

    cfg = cargar_config()
    devices_file = os.path.join(SCRIPT_DIR, cfg["archivos"]["lista_devices"])
    col_ids = cfg["archivos"]["columna_ids"]
    ids = pd.read_csv(devices_file)[col_ids].tolist()

    print(f"=== Análisis Sigfox {month}/{year} — {len(ids)} dispositivos ===")
    registros, dias_mes = recopilar_datos(cfg, ids, year, month)

    if not registros:
        print("No se obtuvo información de ningún dispositivo.")
        sys.exit(1)

    wb = Workbook()
    del wb["Sheet"]  # elimina hoja vacía por defecto
    escribir_hoja_resumen(wb, registros, dias_mes, year, month, cfg, dia_corte)
    escribir_hoja_datos(wb, registros, dias_mes, year, month, cfg)
    escribir_hoja_config(wb, cfg, year, month)

    resultados_dir = os.path.join(SCRIPT_DIR, cfg["archivos"]["directorio_resultados"])
    os.makedirs(resultados_dir, exist_ok=True)
    filename = f"mensajes_{year}_{int(month):02d}.xlsx"
    out_path = os.path.join(resultados_dir, filename)
    wb.save(out_path)
    print(f"\nArchivo guardado: {out_path}")
    print("Listo.")


if __name__ == "__main__":
    main()
